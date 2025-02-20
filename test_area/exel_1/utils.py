from openpyxl.styles import PatternFill, Alignment, Border, Side

SIGNAL_WEEKEND: str = "B"
SIGNAL_WORK: str = "P"
SIGN = Literal[SIGNAL_WEEKEND, SIGNAL_WORK]  # noqa
black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
green_fill = PatternFill(start_color="22e06e", end_color="22e06e", fill_type="solid")
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
# Создаем объект Alignment с выравниванием по центру
center_alignment = Alignment(horizontal="center", vertical="center")
# Создаем объект Border с обрамлением только для левой и правой сторон
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

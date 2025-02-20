from abc import ABC, abstractmethod
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from typing import Literal

SIGNAL_WEEKEND: str = "B"
SIGNAL_WORK: str = "P"
SIGN = Literal[SIGNAL_WEEKEND, SIGNAL_WORK]  # noqa

black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
green_fill = PatternFill(start_color="22e06e", end_color="22e06e", fill_type="solid")
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

center_alignment = Alignment(horizontal="center", vertical="center")
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class IExcelWriter(ABC):
    @abstractmethod
    def save(self):
        pass

    @abstractmethod
    def add_row(self, row: list[str]):
        pass

    @abstractmethod
    def add_cell(self, row: int, column: int, value: str):
        pass


class IExcelStyler(ABC):
    @abstractmethod
    def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
        pass

    @abstractmethod
    def auto_alignment_column_width(self):
        pass

    @abstractmethod
    def auto_alignment_column_center(self):
        pass

    @abstractmethod
    def auto_border(self):
        pass


class Excel(IExcelWriter, IExcelStyler):
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def save(self):
        self.wb.save(self.file)

    def add_row(self, row: list[str]):
        self.sheet.append(row)

    def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
        length = len(colors)
        for index, cell in enumerate(self.sheet[row]):
            if index >= length:
                break
            elif fill := colors[index]:
                cell.fill = fill

    def add_cell(self, row: int, column: int, value: str):
        self.sheet.cell(row=row, column=column, value=value)

    def auto_alignment_column_width(self):
        for columns in self.sheet.iter_cols():
            value = columns[0].value
            length = len(str(value)) + str(value).count(" ") + 2
            column_name = columns[0].column_letter
            self.sheet.column_dimensions[column_name].width = length

    def auto_alignment_column_center(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.alignment = center_alignment

    def auto_border(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.border = border


class IStatistic(ABC):
    @abstractmethod
    def init_date(self, date: str):
        pass

    @property
    @abstractmethod
    def no_driver(self) -> dict[str, int]:
        pass

    @property
    @abstractmethod
    def repair(self) -> dict[str, int]:
        pass

    @property
    @abstractmethod
    def total(self) -> dict[str, int]:
        pass

    @property
    @abstractmethod
    def table(self) -> dict[str, list]:
        pass


class Statistic(IStatistic):
    def __init__(self):
        self._table: dict[str, list] = defaultdict(list)
        self._no_driver = defaultdict(int)
        self._repair = defaultdict(int)
        self._total = defaultdict(int)

    def init_date(self, date: str):
        self._no_driver[date] = 0
        self._repair[date] = 0
        self._total[date] = 0

    @property
    def no_driver(self) -> dict[str, int]:
        return self._no_driver

    @property
    def repair(self) -> dict[str, int]:
        return self._repair

    @property
    def total(self) -> dict[str, int]:
        return self._total

    @property
    def table(self) -> dict[str, list]:
        return self._table


class CrewExel:
    def __init__(
        self,
        file_name: str,
        work_plan: dict[str, dict[str, str]],
        excel_writer: IExcelWriter,
        excel_styler: IExcelStyler,
        statistic: IStatistic,
    ):
        self.file_name = file_name
        self.work_plan = work_plan
        self.excel_writer = excel_writer
        self.excel_styler = excel_styler
        self.statistic = statistic

    def create(self):
        self.excel_writer.add_row(self.titles)

        for name, values in self.table.items():
            self.excel_writer.add_row([name, *values])

        self.excel_writer.add_row([])
        self.any()
        self.excel_writer.save()

    @property
    def titles(self) -> list[str]:
        return ["    Машина    ", *[date for date in self.work_plan.keys()]]

    @property
    def table(self) -> dict[str, list[str]]:
        self._fill_table_data()
        return self.statistic.table

    @property
    def statistic(self) -> IStatistic:
        self._fill_statistic()
        return self.statistic

    def _init_static(self):
        for date in self.work_plan:
            self.statistic.init_date(date)

    def _fill_table_data(self):
        for date, car in self.work_plan.items():
            for name, value in car.items():
                self.statistic.table[name].append(value)

    def _fill_statistic(self):
        self._init_static()
        for date, car in self.work_plan.items():
            for name, value in car.items():
                if value == SIGNAL_WORK:
                    self.statistic.no_driver[date] += 1
                elif value == SIGNAL_WEEKEND:
                    self.statistic.repair[date] += 1
                else:
                    self.statistic.total[date] += 1

    def any(self):
        self.excel_styler.add_color_to_row_cells(
            self.excel_writer.sheet.max_row + 1,
            [black_fill for _ in range(len(self.titles))],
        )
        self.excel_writer.add_row(["без водителя", *self.statistic.no_driver.values()])
        self.fill_color_to_row_cells(list(self.statistic.no_driver.values()))
        self.excel_writer.add_row(["в ремонте", *self.statistic.repair.values()])
        self.fill_color_to_row_cells(list(self.statistic.repair.values()))
        self.excel_writer.add_row(["наряд", *self.statistic.total.values()])
        self.fill_color_to_row_cells(list(self.statistic.total.values()))

    def fill_color_to_row_cells(self, values: list[int]):
        self.excel_styler.add_color_to_row_cells(
            self.excel_writer.sheet.max_row,
            [
                orange_fill,
                *[red_fill if value else green_fill for value in values],
            ],
        )


if __name__ == "__main__":
    from data import data

    excel_writer = Excel("test.xlsx")
    excel_styler = excel_writer
    statistic = Statistic()
    CrewExel("test.xlsx", data, excel_writer, excel_styler, statistic).create()

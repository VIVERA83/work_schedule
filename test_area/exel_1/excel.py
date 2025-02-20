import openpyxl
from openpyxl.styles import PatternFill

from store.excel.utils import (
    center_alignment,
    border,
    black_fill,
    orange_fill,
    red_fill,
    green_fill,
)
from test_area.exel_1.statistic import Statistic


class Excel:
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def save(self):
        self.wb.save(self.file)

    def add_row(self, row: list[str]):
        self.sheet.append(row)

    def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
        length = len(colors)
        for index, cell in enumerate(self.sheet[row]):
            if index >= length:
                break
            elif fill := colors[index]:
                cell.fill = fill

    def add_cell(self, row: int, column: int, value: str):
        self.sheet.cell(row=row, column=column, value=value)

    def auto_alignment_column_width(self):
        for columns in self.sheet.iter_cols():
            value = columns[0].value
            length = len(str(value)) + str(value).count(" ") + 2
            column_name = columns[0].column_letter
            self.sheet.column_dimensions[column_name].width = length

    def auto_alignment_column_center(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.alignment = center_alignment

    def auto_border(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.border = border


class CrewExel:
    def __init__(self, excel: Excel, statistic: Statistic):
        self.excel = excel
        self.statistic = statistic

    def create(self):
        self.excel.add_row(self.statistic.titles)

        for name, values in self.statistic.table.items():
            self.excel.add_row([name, *values])

        self.excel.add_row([])
        self.make_sheet()
        self.excel.save()

    def make_sheet(self):
        self.excel.add_color_to_row_cells(
            self.excel.sheet.max_row + 1,
            [black_fill for _ in range(len(self.statistic.titles))],
        )
        self._add_statistic_row("без водителя", list(self.statistic.no_driver.values()))
        self._add_statistic_row("в ремонте", list(self.statistic.repair.values()))
        self._add_statistic_row("наряд", list(self.statistic.total.values()))

    def _fill_color_to_row_cells(self, values: list[int]):
        self.excel.add_color_to_row_cells(
            self.excel.sheet.max_row,
            [
                orange_fill,
                *[red_fill if value else green_fill for value in values],
            ],
        )

    def _add_statistic_row(self, label: str, values: list[int]):
        self.excel.add_row([label, *values])
        self._fill_color_to_row_cells(values)

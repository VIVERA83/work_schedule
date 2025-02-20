# from store.excel.excel import Excel
# from store.excel.utils import black_fill, orange_fill, red_fill, green_fill
# from store.scheduler.utils import SIGNAL_WORK, SIGNAL_WEEKEND
from typing import Literal
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

SIGNAL_WEEKEND: str = "B"
SIGNAL_WORK: str = "P"
SIGN = Literal[SIGNAL_WEEKEND, SIGNAL_WORK]  # noqa
black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
green_fill = PatternFill(start_color="22e06e", end_color="22e06e", fill_type="solid")
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
# Создаем объект Alignment с выравниванием по центру
center_alignment = Alignment(horizontal="center", vertical="center")
# Создаем объект Border с обрамлением только для левой и правой сторон
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


class Excel:
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def save(self):
        self.wb.save(self.file)

    def add_row(self, row: list[str]):
        self.sheet.append(row)

    def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
        length = len(colors)
        for index, cell in enumerate(self.sheet[row]):
            if index >= length:
                break
            elif fill := colors[index]:
                cell.fill = fill

    def add_cell(self, row: int, column: int, value: str):
        self.sheet.cell(row=row, column=column, value=value)

    def auto_alignment_column_width(self):
        for columns in self.sheet.iter_cols():
            value = columns[0].value
            length = len(str(value)) + str(value).count(" ") + 2
            column_name = columns[0].column_letter
            self.sheet.column_dimensions[column_name].width = length

    def auto_alignment_column_center(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.alignment = center_alignment

    def auto_border(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.border = border


class Statistic:
    def __init__(self):
        self._table: dict[str, list] = defaultdict(list)
        self._no_driver = defaultdict(int)
        self._repair = defaultdict(int)
        self._total = defaultdict(int)
        self._titles: list[str] = []

    @property
    def no_driver(self) -> dict[str, int]:
        return self._no_driver

    @property
    def repair(self) -> dict[str, int]:
        return self._repair

    @property
    def total(self) -> dict[str, int]:
        return self._total

    @property
    def table(self) -> dict[str, list]:
        return self._table

    @property
    def titles(self) -> list[str]:
        return self._titles


class CrewExel:

    def __init__(self, excel: Excel, statistic: Statistic):
        self.excel = excel
        self.statistic = statistic

    def create(self):
        self.excel.add_row(self.statistic.titles)

        for name, values in self.statistic.table.items():
            self.excel.add_row([name, *values])

        self.excel.add_row([])
        self.any()
        self.excel.save()

    # @property
    # def titles(self) -> list[str]:
    #     return ["    Машина    ", *[date for date in self.work_plan.keys()]]
    #
    # @property
    # def table(self) -> dict[str, list[str]]:
    #     self._fill_table_data()
    #     return self._statistic.table
    #
    # @property
    # def statistic(self) -> Statistic:
    #     self._fill_statistic()
    #     return self._statistic
    #
    # def _init_static(self):
    #     for date in self.work_plan:
    #         self._statistic.init_date(date)
    #
    # def _fill_table_data(self):
    #     for date, car in self.work_plan.items():
    #         for name, value in car.items():
    #             self._statistic.table[name].append(value)
    #
    # def _fill_statistic(self):
    #     self._init_static()
    #     for date, car in self.work_plan.items():
    #         for name, value in car.items():
    #             if value == SIGNAL_WORK:
    #                 self._statistic.no_driver[date] += 1
    #             elif value == SIGNAL_WEEKEND:
    #                 self._statistic.repair[date] += 1
    #             else:
    #                 self._statistic.total[date] += 1

    def any(self):
        self.excel.add_color_to_row_cells(self.excel.sheet.max_row + 1, [black_fill for _ in range(len(self.statistic.titles))])
        self.excel.add_row(["без водителя", *self.statistic.no_driver.values()])
        self.fill_color_to_row_cells(list(self.statistic.no_driver.values()))
        self.excel.add_row(["в ремонте", *self.statistic.repair.values()])
        self.fill_color_to_row_cells(list(self.statistic.repair.values()))
        self.excel.add_row(["наряд", *self.statistic.total.values()])
        self.fill_color_to_row_cells(list(self.statistic.total.values()))

    def fill_color_to_row_cells(self, values: list[int]):
        self.excel.add_color_to_row_cells(
            self.excel.sheet.max_row,
            [
                orange_fill,
                *[red_fill if value else green_fill for value in values],
            ],

        )


class StatisticCalculator(Statistic):

    def __init__(self, work_plan: dict[str, dict[str, str]]):
        super().__init__()
        self.work_plan = work_plan
        self._calculate()

    def _init_date(self, date: str):
        self._no_driver[date] = 0
        self._repair[date] = 0
        self._total[date] = 0

    def _init_static(self):
        for date in self.work_plan:
            self._init_date(date)

    def _fill_table_data(self):
        for date, car in self.work_plan.items():
            for name, value in car.items():
                self.table[name].append(value)

    def _fill_statistic(self):
        for date, car in self.work_plan.items():
            for name, value in car.items():
                if value == SIGNAL_WORK:
                    self.no_driver[date] += 1
                elif value == SIGNAL_WEEKEND:
                    self.repair[date] += 1
                else:
                    self.total[date] += 1

    def _fill_titles(self):
        self._titles = ["    Машина    ", *[date for date in self.work_plan.keys()]]

    def _calculate(self):
        self._fill_table_data()
        self._init_static()
        self._fill_statistic()
        self._fill_titles()


class CrewExcelStyler:
    def __init__(self, sheet: Worksheet, statistic: StatisticCalculator):
        self.sheet = sheet
        self.statistic = statistic

    def make_style(self):
        self.add_color_to_row_cells(
            self.sheet.max_row + 1, [black_fill for _ in range(len(self.titles))]
        )
        self.add_row(["без водителя", *self.statistic.no_driver.values()])
        self.fill_color_to_row_cells(list(self.statistic.no_driver.values()))
        self.add_row(["в ремонте", *self.statistic.repair.values()])
        self.fill_color_to_row_cells(list(self.statistic.repair.values()))
        self.add_row(["наряд", *self.statistic.total.values()])
        self.fill_color_to_row_cells(list(self.statistic.total.values()))

    def fill_color_to_row_cells(self, values: list[int]):
        self.add_color_to_row_cells(
            self.sheet.max_row,
            [
                orange_fill,
                *[red_fill if value else green_fill for value in values],
            ],

        )


if __name__ == '__main__':
    from data import data
    excel_ = Excel("test_1.xlsx")
    static = StatisticCalculator(data)
    CrewExel(excel_, static).create()

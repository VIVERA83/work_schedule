import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from excel.utils import (
    border,
    center_alignment,
)


class Excel:
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def get_active_sheet(self) -> Worksheet:
        return self.sheet

    def create_sheet(self, title: str = None) -> Worksheet:
        return self.wb.create_sheet(title)

    def save(self):
        self.wb.save(self.file)

    # # на удаление
    # def add_row(self, row: list[str]):
    #     self.sheet.append(row)
    #
    # # на удаление
    # def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
    #     length = len(colors)
    #     for index, cell in enumerate(self.sheet[row]):
    #         if index >= length:
    #             break
    #         elif fill := colors[index]:
    #             cell.fill = fill
    #
    # # на удаление
    # def add_cell(self, row: int, column: int, value: str):
    #     self.sheet.cell(row=row, column=column, value=value)
    #
    # # на удаление
    # def auto_alignment_column_width(self):
    #     for columns in self.sheet.iter_cols():
    #         value = columns[0].value
    #         length = len(str(value)) + str(value).count(" ") + 2
    #         column_name = columns[0].column_letter
    #         self.sheet.column_dimensions[column_name].width = length
    #
    # # на удаление
    # def auto_alignment_column_center(self):
    #     for columns in self.sheet.iter_cols():
    #         for cell in columns:
    #             cell.alignment = center_alignment
    #
    # # на удаление
    # def auto_border(self):
    #     for columns in self.sheet.iter_cols():
    #         for cell in columns:
    #             cell.border = border

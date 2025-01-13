import openpyxl
from openpyxl.styles import PatternFill

from .utils import center_alignment, border


class Excel:
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def save(self):
        self.wb.save(self.file)

    def add_row(self, row: list[str]):
        self.sheet.append(row)

    def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
        length = len(colors)
        for index, cell in enumerate(self.sheet[row]):
            if index >= length:
                break
            elif fill := colors[index]:
                cell.fill = fill

    def add_cell(self, row: int, column: int, value: str):
        self.sheet.cell(row=row, column=column, value=value)

    def auto_alignment_column_width(self):
        for columns in self.sheet.iter_cols():
            value = columns[0].value
            length = len(str(value)) + str(value).count(" ") + 2
            column_name = columns[0].column_letter
            self.sheet.column_dimensions[column_name].width = length

    def auto_alignment_column_center(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.alignment = center_alignment

    def auto_border(self):
        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.border = border

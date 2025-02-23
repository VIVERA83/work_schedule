from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from excel.sheet import Sheet
from store.excel.dispatchplan import DispatchPlan
from excel.utils import black_fill, orange_fill, red_fill, green_fill


class CrewSheet(Sheet):
    def __init__(self, dispatch_plan: DispatchPlan, sheet: Worksheet):
        super().__init__(sheet)
        self.dispatch_plan = dispatch_plan

    def fill_in_data_sheet(self):
        # добавляем заголовки
        self.add_row(self.dispatch_plan.titles)
        # добавляем данные основные данные распределения наряда
        for name, values in self.dispatch_plan.table.items():
            self.add_row([name, *values])
        # добавляем пустую строку, для разделения
        self.add_row([])
        # добавляем статистику по распределению наряда
        self.add_row(["Без водителя", *list(self.dispatch_plan.no_driver.values())])
        self.add_row(["Ремонт", *list(self.dispatch_plan.repair.values())])
        self.add_row(["Итого в наряде", *list(self.dispatch_plan.total.values())])
        # цветовая заливка
        self._fill_static_row_cells()
        # форматирование строк
        self.auto_border()
        self.auto_alignment_column_width()
        self.auto_alignment_column_center()

    def _fill_static_row_cells(self):
        # номер строка с которой начинается статистика
        start_number_row = len(self.dispatch_plan.table) + 2
        for row, data in enumerate(
            [
                self.dispatch_plan.no_driver.values(),
                self.dispatch_plan.repair.values(),
            ],
            start_number_row + 1,
        ):
            self.add_color_to_row_cells(
                row, [orange_fill, *self._get_row_colors(list(data))]
            )
        # цветовая заливка разделяющая блок наряда и статистики
        self.add_color_to_row_cells(
            start_number_row,
            [black_fill for _ in range(len(self.dispatch_plan.titles))],
        )

    @staticmethod
    def _get_cell_color(status: int) -> PatternFill:
        if status == 0:
            return green_fill
        return red_fill

    def _get_row_colors(self, data: list[int]) -> list[PatternFill]:
        return [self._get_cell_color(value) for value in data]

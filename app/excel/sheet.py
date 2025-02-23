from typing import Any

from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from excel.utils import center_alignment, border


class Sheet:
    def __init__(self, sheet: Worksheet):
        self.sheet = sheet

    def add_cell(self, row: int, column: int, value: str):
        """Добавление значения в ячейку

        :param row: Строка.
        :param column: Столбец.
        :param value: Значение.
        """
        self.sheet.cell(row=row, column=column, value=value)

    def add_row(self, row: list[Any]):
        """Добавление строки в лист.

        :param row: Список значений.
        """
        self.sheet.append(row)

    def add_color_to_row_cells(self, row: int, colors: list[PatternFill]):
        """Добавление цвета в ячейки заданной строки.

        :param row: Строка.
        :param colors: Список цветов.
        """

        length = len(colors)
        for index, cell in enumerate(self.sheet[row]):
            if index >= length:
                break
            elif fill := colors[index]:
                cell.fill = fill

    def auto_alignment_column_width(self):
        """Автоматическое выравнивание ширины столбцов."""

        for columns in self.sheet.iter_cols():
            value = columns[0].value
            length = len(str(value)) + str(value).count(" ") + 2
            column_name = columns[0].column_letter
            self.sheet.column_dimensions[column_name].width = length

    def auto_alignment_column_center(self):
        """Автоматическое выравнивание столбцов по центру."""

        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.alignment = center_alignment

    def auto_border(self):
        """Автоматическое добавление рамки вокруг ячеек."""

        for columns in self.sheet.iter_cols():
            for cell in columns:
                cell.border = border

import openpyxl

from openpyxl.worksheet.worksheet import Worksheet


class Excel:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active

    def get_active_sheet(self) -> Worksheet:
        return self.sheet

    def create_sheet(self, title: str = None) -> Worksheet:
        return self.wb.create_sheet(title)

    def save(self, file_name: str):
        self.wb.save(file_name)
